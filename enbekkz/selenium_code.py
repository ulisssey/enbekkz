from openpyxl import Workbook
from openpyxl.styles import numbers
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.select import Select
from datetime import datetime
from flask import flash
import openpyxl
import pyautogui
import time
import os


login = 'admin.sko'
password = 'Qwerty123@@@'
school = '"ДОД" при №1 Мектеп-гимназия, №1 Мектеп-гимназия'

# Засечение времени
day = datetime.now().day
month = datetime.now().month
year = datetime.now().year

# Сохранение логов
if not os.path.exists('./logs'):
    os.mkdir('./logs')

# Конфигурация драйвера 
executable_path = './chromedriver.exe'
chrome_options = Options()
chrome_options.add_argument('--headless')  # Run Chrome in headless mode
chrome_options.add_argument('--no-sandbox')  # Disable sandbox mode to avoid issues with some systems
chrome_options.add_argument('--disable-dev-shm-usage')  # Disable usage of /dev/shm to avoid issues with some systems

# Set the service_log_path option to a file
service_log_path = 'chromedriver.log'
service_log = open(service_log_path, 'w')
service = webdriver.chrome.service.Service('chromedriver', log_path=service_log_path)
driver = webdriver.Chrome(executable_path=executable_path, chrome_options=chrome_options, service=service)

# Скрипт для логина и перехода настраницу поиска
def main_code():
    driver.get("https://login.kundelik.kz/login")
    driver.maximize_window()
    # Авторизация в login.kundelik.kz
    try:
        if WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, "//input[contains(@class, 'input_login')]"))):
            driver.find_element(By.XPATH, "//input[contains(@class, 'input_login')]").send_keys(login)
            driver.find_element(By.XPATH, "//input[contains(@class, 'input_type-password')]").send_keys(password)
            driver.find_element(By.XPATH, "//input[contains(@class, 'login__submit')]").click()
            try: 
                if WebDriverWait(driver, 5).until(EC.visibility_of_element_located((By.XPATH, "//div[@class='message ']"))):
                    return 'Ошибка при авторизации неправильный логин или пароль'
            except:
                pass
    except:
        pass
    # Путь к странице "Администрирование школы"
    finally:
        WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, "//a[@class='header-logotype header-logotype_kz']")))
        driver.find_element(By.XPATH, "//div[@class='header-localization-select__info']").click()
        WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, "//div[@class='header-localization-select__info header-localization-select__info_active']")))
        driver.find_element(By.XPATH, "//a[@title='Русский']").click()
        WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, "//a[@title='Образование']")))
        driver.find_element(By.XPATH, "//a[@title='Образование']").click()
        WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, "//a[@title='Моя школа']")))
        driver.find_element(By.XPATH, "//a[@title='Моя школа']").click()
        WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, "//a[@id='IDSchoolAdministration']")))
        driver.find_element(By.XPATH, "//a[@id='IDSchoolAdministration']").click()
        WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, "//a[text()='Список людей']")))
        driver.find_element(By.XPATH, "//a[text()='Список людей']").click()
        WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, "//h2[text()='Поиск людей']")))

# Скрипт для получения данных из страницы ученика
def get_data(lang):
    WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, "//dl[@class='info s2']")))
    all_text = driver.find_elements(By.XPATH, "//dl[@class='info s2']/dd")
    fio = all_text[0].text
    gender = all_text[1].text
    if lang == 'kz':
        if gender == 'Мужской':
            gender = 'Ер'
        else:
            gender = 'Әйел'
    birth_date = all_text[3].text
    email = all_text[4].text
    all_text1 = driver.find_elements(By.XPATH, "//dl[@class='info big']/dd")
    school_class = all_text1[0].text
    # Личные данные
    driver.find_element(By.XPATH, "//a[@id='TabPersonal']").click()
    WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, "//label[text()='Фамилия']")))
    iin = driver.find_element(By.XPATH, "//input[@id='personalNumber']").get_attribute('value')
    select = Select(driver.find_element(By.XPATH, "//select[@id='nationality']"))
    nationality = select.first_selected_option.text
    if nationality == '- Выберите национальность -':
        nationality = None
    if lang == 'kz':
        if nationality == 'казах / казашка':
            nationality = 'қазақ'
        elif nationality == 'русский / русская':
            nationality = 'орыс'
    select = Select(driver.find_element(By.XPATH, "//select[@id='nativeLanguage']"))
    native_lg = select.first_selected_option.text
    if native_lg == '- Выберите язык -':
        native_lg = None
    if lang == 'kz':
        if native_lg == 'Казахский':
            native_lg == 'Қазақша'
        elif native_lg == 'Русский':
            native_lg = 'Орысша'
    svidetelstvo = driver.find_element(By.XPATH, "//input[@id='bcert_docnum']").get_attribute('value')
    kem_vydan = driver.find_element(By.XPATH, "//input[@id='bcert_issby']").get_attribute('value')
    data_vydachy = driver.find_element(By.XPATH, "//input[@id='birthcertificatedate']").get_attribute('value')
    mesto_vydachy = driver.find_element(By.XPATH, "//input[@id='bcert_place']").get_attribute('value')
    select = Select(driver.find_element(By.XPATH, "//select[@id='medid']"))
    medid = select.first_selected_option.text
    select = Select(driver.find_element(By.XPATH, "//select[@id='physid']"))
    physid = select.first_selected_option.text
    address = driver.find_element(By.XPATH, "//input[@id='UserControlPersonEdit_Tabs_UserControlPersonEditPersonal_ActualAddress']").get_attribute('value')
    mphone = driver.find_element(By.XPATH, "//input[@id='mphone']").get_attribute('value')
    hphone = driver.find_element(By.XPATH, "//input[@id='hphone']").get_attribute('value')
    # Миграция
    driver.find_element(By.XPATH, "//a[@id='TabMigration']").click()
    WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, "//input[@id='wasInPreschoolInstitution']")))
    pre_school = driver.find_element(By.XPATH, "//input[@id='wasInPreschoolInstitution']").get_attribute('checked') 
    if pre_school == 'checked':
        pre_school = 'Да'
    else:
        pre_school = 'Нет'
    # Родственники
    driver.find_element(By.XPATH, "//a[@id='TabParents']").click()
    WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, "//a[@id='buttonAddRelative']")))
    try:
        if driver.find_element(By.XPATH, "//p[text()='У пользователя не настроены родственные связи']"):
            fio_parent = ''
            work_place = ''
            dolzhnost = ''
            parent_iin = ''
            parent_phone = ''
    except:
        driver.find_element(By.XPATH, "//a[@title='Редактировать']").click()
        all_text = driver.find_elements(By.XPATH, "//dl[@class='info s2']/dd")
        fio_parent = all_text[0].text
        all_text1 = driver.find_elements(By.XPATH, "//dl[@class='info big']/dt")
        all_text2 = driver.find_elements(By.XPATH, "//dl[@class='info big']/dd")
        if all_text1[0].text == 'Должность':
            work_place = school
            dolzhnost = all_text2[0].text
            driver.find_element(By.XPATH, "//a[@id='TabPersonal']").click()
            WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, "//input[@id='personalNumber']")))
            parent_iin = driver.find_element(By.XPATH, "//input[@id='personalNumber']").get_attribute('value')
            parent_phone = driver.find_element(By.XPATH, "//input[@id='mphone']").get_attribute('value')
        else:
            driver.find_element(By.XPATH, "//a[@id='TabPersonal']").click()
            WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, "//input[@id='personalNumber']")))
            parent_iin = driver.find_element(By.XPATH, "//input[@id='personalNumber']").get_attribute('value')
            parent_phone = driver.find_element(By.XPATH, "//input[@id='mphone']").get_attribute('value')
            work_place = driver.find_element(By.XPATH, "//textarea[@id='workplace']").get_attribute('value')
            dolzhnost = driver.find_element(By.XPATH, "//input[@id='workPosition']").get_attribute('value')
        driver.back()
        driver.back()
    return [fio, gender, birth_date, iin, school_class, school_class[:1], mphone, hphone, address, email, nationality, native_lg, svidetelstvo, data_vydachy, kem_vydan, mesto_vydachy, medid, physid, pre_school, fio_parent, work_place, dolzhnost, parent_iin, parent_phone]

# Скрипт для массовой выгрузки
def get_all(lang):
    if lang == 'ru':
        if not os.path.isfile(f'./logs/{day}{month}{year}_Отчет о массовой выгрузке рус.xlsx'):
            wb = Workbook()
            wb.save(f'./logs/{day}{month}{year}_Отчет о массовой выгрузке рус.xlsx')
            wb = openpyxl.load_workbook(f'./logs/{day}{month}{year}_Отчет о массовой выгрузке рус.xlsx')
            ws = wb.active
            ws.append(['Фио', 'Пол', 'Дата рождения', 'ИИН', 'Класс', 'Параллель', 'Моб. телефон', 'Дом. телефон', 'Адрес проживания', 'Эл. почта', 'Национальность', 'Родной язык', 'Свидетельство о рождении', 'Дата выдачи', 'Кем выдан', 'Место выдачи', 'Медицинская группа здоровья', 'Физкультурная группа здоровья', 'Был в дошкольном учреждении', 'Родитель / Законный представитель', 'Место работы', 'Должность', 'ИИН', 'Моб. телефон', 'Ошибки'])
            col = ws.column_dimensions['D']
            col1 = ws.column_dimensions['V']
            col1.number_format = numbers.FORMAT_TEXT
            col.number_format = numbers.FORMAT_TEXT
            for column_cells in ws.columns:
                column_letter = column_cells[0].column_letter
                ws.column_dimensions[column_letter].width = 20
            wb.save(f'./logs/{day}{month}{year}_Отчет о массовой выгрузке рус.xlsx')
    else:
        if not os.path.isfile(f'./logs/{day}{month}{year}_Отчет о массовой выгрузке каз.xlsx'):
            wb = Workbook()
            wb.save(f'./logs/{day}{month}{year}_Отчет о массовой выгрузке каз.xlsx')
            wb = openpyxl.load_workbook(f'./logs/{day}{month}{year}_Отчет о массовой выгрузке каз.xlsx')
            ws = wb.active
            ws.append(['Аты жөні', 'Жынысы', 'Туған күні', 'ЖСН', 'Сынып', 'Параллель', 'Ұялы телефон', 'Үй телефон', 'Тұрғылықты мекенжайы', 'Электрондық пошта', 'Ұлты', 'Ана тілі', 'Туу туралы куәлік', 'Шығарылған күні', 'Кіммен берілді', 'Берілген жері', 'Дәрігерлік денсаулық тобы', 'Дене сауықтыру тобы', 'Мектепке дейінгі мекемеде болды', 'Ата-ана/заңды өкіл', 'Жұмыс орны', 'Лауазымы', 'ЖСН', 'Ұялы телефон', 'Қателер'])
            col = ws.column_dimensions['D']
            col1 = ws.column_dimensions['V']
            col1.number_format = numbers.FORMAT_TEXT
            col.number_format = numbers.FORMAT_TEXT
            for column_cells in ws.columns:
                column_letter = column_cells[0].column_letter
                ws.column_dimensions[column_letter].width = 20
            wb.save(f'./logs/{day}{month}{year}_Отчет о массовой выгрузке каз.xlsx')
    windows = pyautogui.getAllTitles()
    for window in windows:
        if lang == 'kz':
            if f'{day}{month}{year}_Отчет о массовой выгрузке каз' in window:
                return 'excel_error_kz'
                exit()
        elif lang == 'ru':
            if f'{day}{month}{year}_Отчет о массовой выгрузке рус' in window:
                return 'excel_error_ru'
                exit()
    status = main_code()
    if status == 'Ошибка при авторизации неправильный логин или пароль':
        wb = openpyxl.load_workbook(f'./logs/{day}{month}{year}_Отчет о массовой выгрузке.xlsx')
        ws = wb.active
        ws.append(['']*24 + [status])
        wb.save(f'./logs/{day}{month}{year}_Отчет о массовой выгрузке.xlsx')
        driver.quit()
    page_num = driver.find_elements(By.XPATH, "//div[@class='pager']//li")
    num = 2
    pages = int(page_num[-1].text)
    while num <= pages:
        all = driver.find_elements(By.XPATH, "//span[text()='Ученик']/../..//a[@class='u']")
        children = []
        for i in all:
            children.append(i.text)
        for j in children:
            driver.find_element(By.XPATH, f"//a[text()='{j}']").click()
            rows = get_data(lang)
            if lang == 'ru':
                wb = openpyxl.load_workbook(f'./logs/{day}{month}{year}_Отчет о массовой выгрузке рус.xlsx')
                ws = wb.active
                ws.append(rows)
                col = ws.column_dimensions['D']
                col1 = ws.column_dimensions['V']
                col1.number_format = numbers.FORMAT_TEXT
                col.number_format = numbers.FORMAT_TEXT
                wb.save(f'./logs/{day}{month}{year}_Отчет о массовой выгрузке рус.xlsx')
            else:
                wb = openpyxl.load_workbook(f'./logs/{day}{month}{year}_Отчет о массовой выгрузке каз.xlsx')
                ws = wb.active
                ws.append(rows)
                col = ws.column_dimensions['D']
                col1 = ws.column_dimensions['V']
                col1.number_format = numbers.FORMAT_TEXT
                col.number_format = numbers.FORMAT_TEXT
                wb.save(f'./logs/{day}{month}{year}_Отчет о массовой выгрузке каз.xlsx')
            driver.back()
            driver.back()
            driver.back()
            driver.back()
            time.sleep(3)
        driver.find_element(By.XPATH, f"//div[@class='pager']//li/a[text()={num}]").click()
        time.sleep(3)
        num += 1
    if lang == 'kz':
        return 'success_kz'
    else:
        return 'success_ru'

# Скрипт для выгрузки по ФИО
def search_by_fio(fio, lang):
    windows = pyautogui.getAllTitles()
    for window in windows:
        if lang == 'kz':
            if f'{day}{month}{year}_Отчет о выгрузке по ФИО каз' in window:
                return 'excel_error_kz'
                exit()
        elif lang == 'ru':
            if f'{day}{month}{year}_Отчет о выгрузке по ФИО рус' in window:
                return 'excel_error_ru'
                exit()
    main_code()
    driver.find_element(By.XPATH, "//input[@id='search']").send_keys(fio)
    time.sleep(0.8)
    driver.find_element(By.XPATH, "//input[@id='go']").click()
    if driver.find_elements(By.XPATH, "//span[text()='Ученик']/../..//a[@class='u']"):
        if lang == 'ru':
            if not os.path.isfile(f'./logs/{day}{month}{year}_Отчет о выгрузке по ФИО рус.xlsx'):
                wb = Workbook()
                wb.save(f'./logs/{day}{month}{year}_Отчет о выгрузке по ФИО рус.xlsx')
                wb = openpyxl.load_workbook(f'./logs/{day}{month}{year}_Отчет о выгрузке по ФИО рус.xlsx')
                ws = wb.active
                ws.append(['Фио', 'Пол', 'Дата рождения', 'ИИН', 'Класс', 'Параллель', 'Моб. телефон', 'Дом. телефон', 'Адрес проживания', 'Эл. почта', 'Национальность', 'Родной язык', 'Свидетельство о рождении', 'Дата выдачи', 'Кем выдан', 'Место выдачи', 'Медицинская группа здоровья', 'Физкультурная группа здоровья', 'Был в дошкольном учреждении', 'Родитель / Законный представитель', 'Место работы', 'Должность', 'ИИН', 'Моб. телефон', 'Ошибки'])
                col = ws.column_dimensions['D']
                col1 = ws.column_dimensions['V']
                col1.number_format = numbers.FORMAT_TEXT
                col.number_format = numbers.FORMAT_TEXT
                for column_cells in ws.columns:
                    column_letter = column_cells[0].column_letter
                    ws.column_dimensions[column_letter].width = 20
                wb.save(f'./logs/{day}{month}{year}_Отчет о выгрузке по ФИО рус.xlsx')
        else:
            if not os.path.isfile(f'./logs/{day}{month}{year}_Отчет о выгрузке по ФИО каз.xlsx'):
                wb = Workbook()
                wb.save(f'./logs/{day}{month}{year}_Отчет о выгрузке по ФИО каз.xlsx')
                wb = openpyxl.load_workbook(f'./logs/{day}{month}{year}_Отчет о выгрузке по ФИО каз.xlsx')
                ws = wb.active
                ws.append(['Аты жөні', 'Жынысы', 'Туған күні', 'ЖСН', 'Сынып', 'Параллель', 'Ұялы телефон', 'Үй телефон', 'Тұрғылықты мекенжайы', 'Электрондық пошта', 'Ұлты', 'Ана тілі', 'Туу туралы куәлік', 'Шығарылған күні', 'Кіммен берілді', 'Берілген жері', 'Дәрігерлік денсаулық тобы', 'Дене сауықтыру тобы', 'Мектепке дейінгі мекемеде болды', 'Ата-ана/заңды өкіл', 'Жұмыс орны', 'Лауазымы', 'ЖСН', 'Ұялы телефон', 'Қателер'])
                col = ws.column_dimensions['D']
                col1 = ws.column_dimensions['V']
                col1.number_format = numbers.FORMAT_TEXT
                col.number_format = numbers.FORMAT_TEXT
                for column_cells in ws.columns:
                    column_letter = column_cells[0].column_letter
                    ws.column_dimensions[column_letter].width = 20
                wb.save(f'./logs/{day}{month}{year}_Отчет о выгрузке по ФИО каз.xlsx')
        all = driver.find_elements(By.XPATH, "//span[text()='Ученик']/../..//a[@class='u']")
        children = []
        for i in all:
            children.append(i.text)
        for j in children:
            driver.find_element(By.XPATH, f"//a[text()='{j}']").click()
            rows = get_data(lang)
            if lang == 'ru':
                wb = openpyxl.load_workbook(f'./logs/{day}{month}{year}_Отчет о выгрузке по ФИО рус.xlsx')
                ws = wb.active
                ws.append(rows)
                col = ws.column_dimensions['D']
                col1 = ws.column_dimensions['V']
                col1.number_format = numbers.FORMAT_TEXT
                col.number_format = numbers.FORMAT_TEXT
                wb.save(f'./logs/{day}{month}{year}_Отчет о выгрузке по ФИО рус.xlsx')
            else:
                wb = openpyxl.load_workbook(f'./logs/{day}{month}{year}_Отчет о выгрузке по ФИО каз.xlsx')
                ws = wb.active
                ws.append(rows)
                col = ws.column_dimensions['D']
                col1 = ws.column_dimensions['V']
                col1.number_format = numbers.FORMAT_TEXT
                col.number_format = numbers.FORMAT_TEXT
                wb.save(f'./logs/{day}{month}{year}_Отчет о выгрузке по ФИО каз.xlsx')
            driver.back()
            driver.back()
            driver.back()
            driver.back()
            time.sleep(3)
    else:
        if lang == 'ru':
        #     wb = openpyxl.load_workbook(f'./logs/{day}{month}{year}_Отчет о выгрузке по ФИО рус.xlsx')
        #     ws = wb.active
        #     ws.append([fio] + ['']*23 + ['Не найден ученик с таким ФИО'])
        #     wb.save(f'./logs/{day}{month}{year}_Отчет о выгрузке по ФИО рус.xlsx')
            return fio
        else:
            # wb = openpyxl.load_workbook(f'./logs/{day}{month}{year}_Отчет о выгрузке по ФИО каз.xlsx')
            # ws = wb.active
            # ws.append([fio] + ['']*23 + ['Мұндай аты-жөнімен оқушы табылмады'])
            # wb.save(f'./logs/{day}{month}{year}_Отчет о выгрузке по ФИО каз.xlsx')
            return fio

# Скрипт для выгрузки по классам
def search_by_class(class_text, lang):
    windows = pyautogui.getAllTitles()
    for window in windows:
        if lang == 'kz':
            if f'{day}{month}{year}_Отчет о выгрузке по классу каз' in window:
                return 'excel_error_kz'
                exit()
        elif lang == 'ru':
            if f'{day}{month}{year}_Отчет о выгрузке по классу рус' in window:
                return 'excel_error_ru'
                exit()
    main_code()
    driver.find_element(By.XPATH, "//input[@id='class']").send_keys(class_text)
    time.sleep(0.8)
    driver.find_element(By.XPATH, "//input[@id='go']").click()
    time.sleep(5)
    if driver.find_elements(By.XPATH, "//div[@class='pager']//li"):
        page_num = driver.find_elements(By.XPATH, "//div[@class='pager']//li")
        num = 2
        pages = int(page_num[-1].text)
        all_pupil = []
        while num <= pages:
            all = driver.find_elements(By.XPATH, "//span[text()='Ученик']/../..//a[@class='u']")
            if not all:
                pass
            else:
                if lang == 'ru':
                    if not os.path.isfile(f'./logs/{day}{month}{year}_Отчет о выгрузке по классу рус.xlsx'):
                        wb = Workbook()
                        wb.save(f'./logs/{day}{month}{year}_Отчет о выгрузке по классу рус.xlsx')
                        wb = openpyxl.load_workbook(f'./logs/{day}{month}{year}_Отчет о выгрузке по классу рус.xlsx')
                        ws = wb.active
                        ws.append(['Фио', 'Пол', 'Дата рождения', 'ИИН', 'Класс', 'Параллель', 'Моб. телефон', 'Дом. телефон', 'Адрес проживания', 'Эл. почта', 'Национальность', 'Родной язык', 'Свидетельство о рождении', 'Дата выдачи', 'Кем выдан', 'Место выдачи', 'Медицинская группа здоровья', 'Физкультурная группа здоровья', 'Был в дошкольном учреждении', 'Родитель / Законный представитель', 'Место работы', 'Должность', 'ИИН', 'Моб. телефон', 'Ошибки'])
                        for column_cells in ws.columns:
                            column_letter = column_cells[0].column_letter
                            ws.column_dimensions[column_letter].width = 20
                        wb.save(f'./logs/{day}{month}{year}_Отчет о выгрузке по классу рус.xlsx')
                else:
                    if not os.path.isfile(f'./logs/{day}{month}{year}_Отчет о выгрузке по классу каз.xlsx'):
                        wb = Workbook()
                        wb.save(f'./logs/{day}{month}{year}_Отчет о выгрузке по классу каз.xlsx')
                        wb = openpyxl.load_workbook(f'./logs/{day}{month}{year}_Отчет о выгрузке по классу каз.xlsx')
                        ws = wb.active
                        ws.append(['Аты жөні', 'Жынысы', 'Туған күні', 'ЖСН', 'Сынып', 'Параллель', 'Ұялы телефон', 'Үй телефон', 'Тұрғылықты мекенжайы', 'Электрондық пошта', 'Ұлты', 'Ана тілі', 'Туу туралы куәлік', 'Шығарылған күні', 'Кіммен берілді', 'Берілген жері', 'Дәрігерлік денсаулық тобы', 'Дене сауықтыру тобы', 'Мектепке дейінгі мекемеде болды', 'Ата-ана/заңды өкіл', 'Жұмыс орны', 'Лауазымы', 'ЖСН', 'Ұялы телефон', 'Қателер'])
                        for column_cells in ws.columns:
                            column_letter = column_cells[0].column_letter
                            ws.column_dimensions[column_letter].width = 20
                        wb.save(f'./logs/{day}{month}{year}_Отчет о выгрузке по классу каз.xlsx')
                for i in all:
                    all_pupil.append(i.text)
                children = []
                for i in all:
                    children.append(i.text)
                for j in children:
                    driver.find_element(By.XPATH, f"//a[text()='{j}']").click()
                    rows = get_data(lang)
                    if lang == 'ru':
                        wb = openpyxl.load_workbook(f'./logs/{day}{month}{year}_Отчет о выгрузке по классу рус.xlsx')
                        ws = wb.active
                        ws.append(rows)
                        col = ws.column_dimensions['D']
                        col1 = ws.column_dimensions['V']
                        col1.number_format = numbers.FORMAT_TEXT
                        col.number_format = numbers.FORMAT_TEXT
                        wb.save(f'./logs/{day}{month}{year}_Отчет о выгрузке по классу рус.xlsx')
                    else:
                        wb = openpyxl.load_workbook(f'./logs/{day}{month}{year}_Отчет о выгрузке по классу каз.xlsx')
                        ws = wb.active
                        ws.append(rows)
                        col = ws.column_dimensions['D']
                        col1 = ws.column_dimensions['V']
                        col1.number_format = numbers.FORMAT_TEXT
                        col.number_format = numbers.FORMAT_TEXT
                        wb.save(f'./logs/{day}{month}{year}_Отчет о выгрузке по классу каз.xlsx')
                    driver.back()
                    driver.back()
                    driver.back()
                    driver.back()
                    time.sleep(3)
            driver.find_element(By.XPATH, f"//div[@class='pager']//li/a[text()={num}]").click()
            time.sleep(3)
            num += 1
        if len(all_pupil) == 0:
            if lang == 'kz':
                return 'class_error_kz'
            else:
                return 'class_error_ru'
        else:
            all_pupil.clear()
    else:
        all = driver.find_elements(By.XPATH, "//span[text()='Ученик']/../..//a[@class='u']")
        if not all:
            return class_text
        else:
            if lang == 'ru':
                if not os.path.isfile(f'./logs/{day}{month}{year}_Отчет о выгрузке по классу рус.xlsx'):
                    wb = Workbook()
                    wb.save(f'./logs/{day}{month}{year}_Отчет о выгрузке по классу рус.xlsx')
                    wb = openpyxl.load_workbook(f'./logs/{day}{month}{year}_Отчет о выгрузке по классу рус.xlsx')
                    ws = wb.active
                    ws.append(['Фио', 'Пол', 'Дата рождения', 'ИИН', 'Класс', 'Параллель', 'Моб. телефон', 'Дом. телефон', 'Адрес проживания', 'Эл. почта', 'Национальность', 'Родной язык', 'Свидетельство о рождении', 'Дата выдачи', 'Кем выдан', 'Место выдачи', 'Медицинская группа здоровья', 'Физкультурная группа здоровья', 'Был в дошкольном учреждении', 'Родитель / Законный представитель', 'Место работы', 'Должность', 'ИИН', 'Моб. телефон', 'Ошибки'])
                    col = ws.column_dimensions['D']
                    col1 = ws.column_dimensions['V']
                    col1.number_format = numbers.FORMAT_TEXT
                    col.number_format = numbers.FORMAT_TEXT
                    for column_cells in ws.columns:
                        column_letter = column_cells[0].column_letter
                        ws.column_dimensions[column_letter].width = 20
                    wb.save(f'./logs/{day}{month}{year}_Отчет о выгрузке по классу рус.xlsx')
            else:
                if not os.path.isfile(f'./logs/{day}{month}{year}_Отчет о выгрузке по классу каз.xlsx'):
                    wb = Workbook()
                    wb.save(f'./logs/{day}{month}{year}_Отчет о выгрузке по классу каз.xlsx')
                    wb = openpyxl.load_workbook(f'./logs/{day}{month}{year}_Отчет о выгрузке по классу каз.xlsx')
                    ws = wb.active
                    ws.append(['Аты жөні', 'Жынысы', 'Туған күні', 'ЖСН', 'Сынып', 'Параллель', 'Ұялы телефон', 'Үй телефон', 'Тұрғылықты мекенжайы', 'Электрондық пошта', 'Ұлты', 'Ана тілі', 'Туу туралы куәлік', 'Шығарылған күні', 'Кіммен берілді', 'Берілген жері', 'Дәрігерлік денсаулық тобы', 'Дене сауықтыру тобы', 'Мектепке дейінгі мекемеде болды', 'Ата-ана/заңды өкіл', 'Жұмыс орны', 'Лауазымы', 'ЖСН', 'Ұялы телефон', 'Қателер'])
                    col = ws.column_dimensions['D']
                    col1 = ws.column_dimensions['V']
                    col1.number_format = numbers.FORMAT_TEXT
                    col.number_format = numbers.FORMAT_TEXT
                    for column_cells in ws.columns:
                        column_letter = column_cells[0].column_letter
                        ws.column_dimensions[column_letter].width = 20
                    wb.save(f'./logs/{day}{month}{year}_Отчет о выгрузке по классу каз.xlsx')
            children = []
            for i in all:
                children.append(i.text)
            for j in children:
                driver.find_element(By.XPATH, f"//a[text()='{j}']").click()
                rows = get_data(lang)
                if lang == 'ru':
                    wb = openpyxl.load_workbook(f'./logs/{day}{month}{year}_Отчет о выгрузке по классу рус.xlsx')
                    ws = wb.active
                    ws.append(rows)
                    col = ws.column_dimensions['D']
                    col1 = ws.column_dimensions['V']
                    col1.number_format = numbers.FORMAT_TEXT
                    col.number_format = numbers.FORMAT_TEXT
                    wb.save(f'./logs/{day}{month}{year}_Отчет о выгрузке по классу рус.xlsx')
                else:
                    wb = openpyxl.load_workbook(f'./logs/{day}{month}{year}_Отчет о выгрузке по классу каз.xlsx')
                    ws = wb.active
                    ws.append(rows)
                    col = ws.column_dimensions['D']
                    col1 = ws.column_dimensions['V']
                    col1.number_format = numbers.FORMAT_TEXT
                    col.number_format = numbers.FORMAT_TEXT
                    wb.save(f'./logs/{day}{month}{year}_Отчет о выгрузке по классу каз.xlsx')
                driver.back()
                driver.back()
                driver.back()
                driver.back()
                time.sleep(3)
